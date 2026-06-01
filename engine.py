import pandas as pd
import re
from datetime import datetime, timedelta, time
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SKIP_STATUSES = ("OFF","Eğitim","İAF","Rapor","Ücretsiz İzin","Taziye İzni","CHAT","CALL","Yıllık İzin")
START_TOL    = pd.Timedelta(seconds=10)
DISC_TOL     = pd.Timedelta(seconds=10)
DISC_MAX_DUR = 30
GECIS_LIMIT  = 600

COLORS = {
    "Geç Giriş":"FFF2CC","Erken Çıkış":"FCE4D6","Fazla Short Break":"DEEBF7",
    "Fazla Lunch":"E2EFDA","Fazla Geçiş Kullanımı":"F4CCFF",
    "Uzun Geçiş Süresi":"FFE0CC","Hatalı Statü (Invisible)":"FFD0D0",
}
IHLAL_TYPES_CHAT = ["Geç Giriş","Erken Çıkış","Fazla Short Break","Fazla Lunch",
                    "Fazla Geçiş Kullanımı","Uzun Geçiş Süresi","Hatalı Statü (Invisible)"]
IHLAL_TYPES_CALL = ["Geç Giriş","Erken Çıkış","Fazla Short Break","Fazla Lunch"]

def fmt(sn):
    sn=int(sn)
    if sn<=0: return "-"
    if sn>=3600: return f"{sn//3600}sa {(sn%3600)//60}dk {sn%60}sn"
    return f"{sn//60}dk {sn%60}sn"

# ── Vardiya parse ─────────────────────────────────────────────────────────────
def parse_shift(s, sd):
    if not isinstance(s, str): return None
    s=s.strip()
    if s in SKIP_STATUSES or s=="": return None
    paren=re.search(r"\((.+?)\)",s); plain=re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})",s)
    if not plain: return None
    bs,be=plain.group(1),plain.group(2)
    if paren:
        inner=paren.group(1).strip(); full=re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})",inner)
        if full: end_str=full.group(2)
        else:
            single=re.match(r"(\d{2}:\d{2})",inner)
            end_str=single.group(1) if single else be
        start_str=bs
    else: start_str,end_str=bs,be
    sh,sm=map(int,start_str.split(":")); eh,em=map(int,end_str.split(":"))
    sdt=datetime.combine(sd.date(),time(sh,sm)); edt=datetime.combine(sd.date(),time(eh,em))
    if edt<=sdt: edt+=timedelta(days=1)
    return (sdt,edt)

def build_schedules(df_v, agent_col, date_cols, dcf, name_map=None):
    if name_map is None: name_map={}
    def gsn(v): return name_map.get(str(v).strip().lower(),str(v).strip())
    ss={}; bd={}
    for _,row in df_v.iterrows():
        ag=gsn(row[agent_col])
        for col in date_cols:
            cf=dcf[col]; val=row[col]; r=parse_shift(val,cf)
            if r: ss[(ag,cf.date())]=r
            elif isinstance(val,str) and val.strip() in SKIP_STATUSES: bd[(ag,cf.date())]=True
    return ss,bd

# ── Call penceresi ────────────────────────────────────────────────────────────
def build_call_windows(df_call):
    df_call=df_call.copy()
    df_call["Tarih"]=pd.to_datetime(df_call["Tarih"],errors="coerce")
    df_call=df_call.sort_values(["agent","Tarih"])
    windows={}
    for agent,ag_df in df_call.groupby("agent"):
        logins=ag_df[ag_df["Statu"]=="login"]["Tarih"].sort_values().tolist()
        logoffs=ag_df[ag_df["Statu"].isin(["logoff","logoffstart"])]["Tarih"].sort_values().tolist()
        sessions=[]; used=set()
        for lg in logins:
            for i,lo in enumerate(logoffs):
                if lo>lg and i not in used: sessions.append((lg,lo)); used.add(i); break
        if sessions: windows[agent]=sessions
    return windows

def in_call_window(ts,agent,cw): return any(s<=ts<=e for s,e in cw.get(agent,[]))
def before_first_call(ts,agent,cw):
    wins=cw.get(agent,[])
    if not wins: return False
    return ts<min(w[0] for w in wins)
def has_call_transition(agent,shift_start,shift_end,cw):
    for (cs,ce) in cw.get(agent,[]):
        if max(cs,shift_start)<min(ce,shift_end): return True
    return False

def is_in_any_shift(ts,agent,ss):
    return any(ag==agent and s<=ts<=e for (ag,_),(s,e) in ss.items())
def is_sess_in_shift(s,e,agent,ss):
    return any(ag==agent and max(0,(min(e,she)-max(s,shs)).total_seconds())>0 for (ag,_),(shs,she) in ss.items())
def is_blocked(ts,agent,bd,ss):
    if not bd.get((agent,ts.date()),False): return False
    return not is_in_any_shift(ts,agent,ss)

def calc_tamamlatilan_call(agent,ss,bd,df_st,missing_sn):
    if missing_sn<=0: return 0
    ag_df=df_st[df_st["agent"]==agent].sort_values("Tarih")
    logins=ag_df[ag_df["Statu"]=="login"]["Tarih"].sort_values().tolist()
    logoffs=ag_df[ag_df["Statu"]=="logoff"]["Tarih"].sort_values().tolist()
    sessions=[]; used=set()
    for lg in logins:
        for i,lo in enumerate(logoffs):
            if lo>lg and i not in used: sessions.append((lg,lo)); used.add(i); break
    total=0
    for s,e in sessions:
        if is_sess_in_shift(s,e,agent,ss): continue
        if is_blocked(s,agent,bd,ss) or is_blocked(e,agent,bd,ss): continue
        total+=(e-s).total_seconds()
    return int(min(missing_sn,total))

def calc_tamamlatilan_chat(agent,ss,bd,df_st,missing_sn):
    if missing_sn<=0: return 0
    rows=df_st[(df_st["agent"]==agent)&(df_st["statu"]=="Unified online")]
    total=0
    for _,r in rows.iterrows():
        ts=r["start_ts"]; dur=r["sure"]
        if pd.isna(ts) or dur<=0: continue
        te=ts+timedelta(seconds=dur)
        if is_sess_in_shift(ts,te,agent,ss): continue
        if is_blocked(ts,agent,bd,ss) or is_blocked(te,agent,bd,ss): continue
        total+=dur
    return int(min(missing_sn,total))

def calc_login_call(agent,df_call,ss):
    ag_df=df_call[df_call["agent"]==agent].sort_values("Tarih")
    logins=ag_df[ag_df["Statu"]=="login"]["Tarih"].sort_values().tolist()
    logoffs=ag_df[ag_df["Statu"]=="logoff"]["Tarih"].sort_values().tolist()
    sessions=[]; used=set()
    for lg in logins:
        for i,lo in enumerate(logoffs):
            if lo>lg and i not in used: sessions.append((lg,lo)); used.add(i); break
    total=0
    for (a,sd),(sh_s,sh_e) in ss.items():
        if a!=agent: continue
        for (cs,ce) in sessions:
            total+=max(0,(min(ce,sh_e)-max(cs,sh_s)).total_seconds())
    return int(total)

def calc_login_chat(agent,df_st,ss):
    ag_df=df_st[(df_st["agent"]==agent)&(df_st["statu"]=="Unified online")]
    total=0
    for (a,sd),(sh_s,sh_e) in ss.items():
        if a!=agent: continue
        sess=ag_df[(ag_df["start_ts"]>=sh_s)&(ag_df["start_ts"]<=sh_e)]
        total+=sess["sure"].sum()
    return int(total)

def mola_detay_str(rows,label):
    parts=[]
    for _,r in rows.iterrows():
        s=r["start_ts"].strftime("%H:%M")
        e=r["end_ts"].strftime("%H:%M") if pd.notna(r["end_ts"]) else "?"
        sn=int(r["sure"]); parts.append(f"{s}-{e} ({sn//60}dk {sn%60}sn)")
    return f" | {label}: "+", ".join(parts) if parts else ""

def find_system_invisibles(invis_rows,session):
    disc_times=session[session["statu"]=="Disconnected"]["start_ts"].tolist()
    disc_ends=session[session["statu"]=="Disconnected"]["end_ts"].dropna().tolist()
    all_disc=disc_times+disc_ends
    sistem=[]; diger=[]
    for _,ir in invis_rows.iterrows():
        is_s=(ir["sure"]<DISC_MAX_DUR and any(abs(ir["start_ts"]-d)<=DISC_TOL for d in all_disc))
        (sistem if is_s else diger).append(ir)
    diger_df=pd.DataFrame(diger) if diger else pd.DataFrame(columns=invis_rows.columns)
    return sistem,diger_df

def merge_gecis_invisible(gecis_rows,invis_rows):
    if gecis_rows.empty and invis_rows.empty: return [],[]
    invis_list=list(invis_rows.iterrows()) if not invis_rows.empty else []
    gecis_list=list(gecis_rows.iterrows()) if not gecis_rows.empty else []
    matched=set(); tekil=[]
    for gi,gr in gecis_list:
        found=None
        for ii,ir in invis_list:
            if ii in matched: continue
            sm=abs(gr["start_ts"]-ir["start_ts"])<=START_TOL
            em=(not pd.isna(gr["end_ts"]) and not pd.isna(ir["end_ts"]) and gr["end_ts"]==ir["end_ts"])
            if sm or em: found=(ii,ir); break
        if found:
            ii,ir=found; matched.add(ii)
            tekil.append({"start_ts":gr["start_ts"],"end_ts":gr["end_ts"],
                          "sure":max(gr["sure"],ir["sure"]),"label":"Mola/Yemek/Çıkış Geçiş"})
        else:
            tekil.append({"start_ts":gr["start_ts"],"end_ts":gr["end_ts"],
                          "sure":gr["sure"],"label":"Mola/Yemek/Çıkış Geçiş"})
    unmatched=[ir for ii,ir in invis_list if ii not in matched]
    unm_s=sorted(unmatched,key=lambda r:r["start_ts"])
    se=[]; ss2=[]; di=[]
    for ir in unm_s:
        e=ir["end_ts"]; s=ir["start_ts"]
        if not pd.isna(e) and e in se: continue
        if any(abs(s-x)<=START_TOL for x in ss2): continue
        di.append(ir)
        if not pd.isna(e): se.append(e); ss2.append(s)
    tek_s=sorted(tekil,key=lambda r:r["start_ts"])
    se2=[]; ss3=[]; dg=[]
    for gr in tek_s:
        e=gr["end_ts"]; s=gr["start_ts"]
        if not pd.isna(e) and e in se2: continue
        if any(abs(s-x)<=START_TOL for x in ss3): continue
        dg.append(gr)
        if not pd.isna(e): se2.append(e); ss3.append(s)
    return dg,di

# ── CALL ANALİZ ───────────────────────────────────────────────────────────────
def analyze_call(df_status, df_vardiya, agent_col, date_cols, dcf,
                 df_chat_status=None, name_map=None):
    df_status=df_status.copy()
    df_status["agent"]=df_status["Musteri_Temsilcisi"].str.strip()
    if name_map:
        df_status["agent"]=df_status["agent"].apply(lambda v: name_map.get(v.lower(),v))

    ss,bd=build_schedules(df_vardiya,agent_col,date_cols,dcf,name_map)

    chat_inv_cache={}
    if df_chat_status is not None:
        df_cs2=df_chat_status.copy()
        df_cs2["agent"]=df_cs2["Temsilci adı"].ffill().str.strip()
        if name_map: df_cs2["agent"]=df_cs2["agent"].apply(lambda v: name_map.get(v.lower(),v))
        df_cs2["statu"]=df_cs2["Hazır bulunma"].str.strip()
        df_cs2["start_ts"]=pd.to_datetime(df_cs2["Bulunma başlangıç zamanı - Zaman damgası"],errors="coerce")
        df_cs2["sure"]=pd.to_numeric(df_cs2["Temsilcinin bulunduğu süre/saniye"],errors="coerce").fillna(0)
        for ag in df_status["agent"].unique():
            inv=df_cs2[(df_cs2["agent"]==ag)&(df_cs2["statu"]=="Invisible")]
            wins=[]
            for _,r in inv.iterrows():
                if not pd.isna(r["start_ts"]): wins.append((r["start_ts"],r["start_ts"]+timedelta(seconds=r["sure"])))
            chat_inv_cache[ag]=wins

    def in_chat_inv(ts,agent): return any(s<=ts<=e for s,e in chat_inv_cache.get(agent,[]))

    violations=[]
    for (agent,shift_date),(shift_start,shift_end) in ss.items():
        ag_df=df_status[df_status["agent"]==agent].copy()
        mask=(ag_df["Tarih"]>=shift_start-timedelta(hours=1))&(ag_df["Tarih"]<=shift_end+timedelta(hours=1))
        session=ag_df[mask].sort_values("Tarih")
        if session.empty: continue
        logins=session[session["Statu"]=="login"]["Tarih"].sort_values().tolist()
        logoffs=session[session["Statu"]=="logoff"]["Tarih"].sort_values().tolist()
        if not logins: continue
        first_login=logins[0]; last_logoff=logoffs[-1] if logoffs else None

        late=(first_login-shift_start).total_seconds()
        if late>60 and not in_chat_inv(first_login,agent):
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Geç Giriş",
                "Detay":f"Vardiya: {shift_start.strftime('%H:%M')}, İlk Login: {first_login.strftime('%H:%M:%S')} (+{int(late//60)}dk {int(late%60)}sn)",
                "İhlal Süresi (sn)":int(late)})

        if last_logoff is not None:
            early=(shift_end-last_logoff).total_seconds()
            if early>60 and not any(lg>last_logoff for lg in logins) and not in_chat_inv(last_logoff,agent):
                violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                    "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                    "İhlal Türü":"Erken Çıkış",
                    "Detay":f"Vardiya: {shift_end.strftime('%H:%M')}, Son Logoff: {last_logoff.strftime('%H:%M:%S')} (-{int(early//60)}dk {int(early%60)}sn)",
                    "İhlal Süresi (sn)":int(early)})

        shift_session=session[(session["Tarih"]>=shift_start)&(session["Tarih"]<=shift_end)]
        total_sb=shift_session[shift_session["Statu"]=="shortbreak"]["Sure"].sum()
        total_ln=shift_session[shift_session["Statu"]=="lunch"]["Sure"].sum()
        if total_sb>1800:
            ex=total_sb-1800
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Fazla Short Break",
                "Detay":f"Kullanılan: {int(total_sb//60)}dk {int(total_sb%60)}sn / Limit: 30dk / Fazla: {int(ex//60)}dk {int(ex%60)}sn",
                "İhlal Süresi (sn)":int(ex)})
        if total_ln>1800:
            ex=total_ln-1800
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Fazla Lunch",
                "Detay":f"Kullanılan: {int(total_ln//60)}dk {int(total_ln%60)}sn / Limit: 30dk / Fazla: {int(ex//60)}dk {int(ex%60)}sn",
                "İhlal Süresi (sn)":int(ex)})

    df_v=pd.DataFrame(violations).sort_values(["Temsilci","Tarih"]) if violations else \
         pd.DataFrame(columns=["Temsilci","Tarih","Vardiya","İhlal Türü","Detay","İhlal Süresi (sn)"])
    ew={}
    for agent in df_v["Temsilci"].unique():
        ag_df=df_v[df_v["Temsilci"]==agent]
        msn=ag_df[ag_df["İhlal Türü"].isin(["Geç Giriş","Erken Çıkış"])]["İhlal Süresi (sn)"].sum()
        ew[agent]=calc_tamamlatilan_call(agent,ss,bd,df_status,msn)
    login_data={}
    for ag in sorted(set(k[0] for k in ss.keys())):
        login_data[ag]=calc_login_call(ag,df_status,ss)
    return df_v,ew,login_data,ss

# ── CHAT ANALİZ ───────────────────────────────────────────────────────────────
def analyze_chat(df_cs, df_vardiya, agent_col, date_cols, dcf,
                 df_call=None, break_limit=1800, lunch_limit=1800, name_map=None):
    df_cs=df_cs.copy()
    df_cs["agent"]=df_cs["Temsilci adı"].ffill().str.strip()
    df_cs["statu"]=df_cs["Hazır bulunma"].str.strip()
    df_cs["sure"]=pd.to_numeric(df_cs["Temsilcinin bulunduğu süre/saniye"],errors="coerce").fillna(0)
    df_cs["start_ts"]=pd.to_datetime(df_cs["Bulunma başlangıç zamanı - Zaman damgası"],errors="coerce")
    df_cs["end_ts"]=pd.to_datetime(df_cs["Bulunma bitiş zamanı - Zaman damgası"],errors="coerce")
    if name_map: df_cs["agent"]=df_cs["agent"].apply(lambda v: name_map.get(v.lower(),v))

    cw={}
    if df_call is not None:
        df_call=df_call.copy()
        df_call["agent"]=df_call["Musteri_Temsilcisi"].str.strip()
        cw=build_call_windows(df_call)

    if name_map:
        df_vardiya=df_vardiya.copy()
        df_vardiya[agent_col]=df_vardiya[agent_col].apply(lambda v: name_map.get(str(v).strip().lower(),str(v).strip()))
    ss,bd=build_schedules(df_vardiya,agent_col,date_cols,dcf,name_map)

    def in_cw(t,ag): return in_call_window(t,ag,cw)
    def bfc(t,ag): return before_first_call(t,ag,cw)
    def filter_call(df_rows,agent):
        if df_rows.empty: return df_rows
        return df_rows[~df_rows["start_ts"].apply(lambda t: in_cw(t,agent))].copy()

    violations=[]; sistem_kes={}

    for (agent,shift_date),(shift_start,shift_end) in ss.items():
        ag_df=df_cs[df_cs["agent"]==agent].copy()
        mask=(ag_df["start_ts"]>=shift_start-timedelta(hours=1))&(ag_df["start_ts"]<=shift_end+timedelta(hours=1))
        session=ag_df[mask].sort_values("start_ts").reset_index(drop=True)
        if session.empty: continue
        online_rows=session[session["statu"]=="Unified online"]["start_ts"].sort_values().tolist()
        offline_rows=session[session["statu"].isin(["Unified offline","Disconnected"])]["start_ts"].sort_values().tolist()
        if not online_rows: continue
        first_online=online_rows[0]; last_offline=offline_rows[-1] if offline_rows else None

        late=(first_online-shift_start).total_seconds()
        if late>60 and not (in_cw(first_online,agent) or bfc(first_online,agent)):
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Geç Giriş",
                "Detay":f"Vardiya: {shift_start.strftime('%H:%M')}, İlk Online: {first_online.strftime('%H:%M:%S')} (+{int(late//60)}dk {int(late%60)}sn)",
                "İhlal Süresi (sn)":int(late)})

        if last_offline is not None:
            early=(shift_end-last_offline).total_seconds()
            if early>60 and not any(on>last_offline for on in online_rows) and not in_cw(last_offline,agent):
                violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                    "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                    "İhlal Türü":"Erken Çıkış",
                    "Detay":f"Vardiya: {shift_end.strftime('%H:%M')}, Son Çıkış: {last_offline.strftime('%H:%M:%S')} (-{int(early//60)}dk {int(early%60)}sn)",
                    "İhlal Süresi (sn)":int(early)})

        ss_df=session[(session["start_ts"]>=shift_start)&(session["start_ts"]<=shift_end)].copy()
        sb_f=filter_call(ss_df[ss_df["statu"]=="Mola"],agent)
        ln_f=filter_call(ss_df[ss_df["statu"]=="Yemek"],agent)
        total_sb=sb_f["sure"].sum() if not sb_f.empty else 0
        total_ln=ln_f["sure"].sum() if not ln_f.empty else 0

        if total_sb>break_limit:
            ex=total_sb-break_limit
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Fazla Short Break",
                "Detay":(f"Kullanılan: {int(total_sb//60)}dk {int(total_sb%60)}sn / Limit: {break_limit//60}dk / Fazla: {int(ex//60)}dk {int(ex%60)}sn"
                         +mola_detay_str(sb_f,"Molalar")),
                "İhlal Süresi (sn)":int(ex)})
        if total_ln>lunch_limit:
            ex=total_ln-lunch_limit
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Fazla Lunch",
                "Detay":(f"Kullanılan: {int(total_ln//60)}dk {int(total_ln%60)}sn / Limit: {lunch_limit//60}dk / Fazla: {int(ex//60)}dk {int(ex%60)}sn"
                         +mola_detay_str(ln_f,"Yemekler")),
                "İhlal Süresi (sn)":int(ex)})

        gecis_rows=ss_df[ss_df["statu"]=="Mola/Yemek/Çıkış Geçiş"].copy()
        invis_rows=ss_df[ss_df["statu"]=="Invisible"].copy()
        sistem_invis,invis_clean=find_system_invisibles(invis_rows,ss_df)
        if sistem_invis:
            sistem_kes[agent]=sistem_kes.get(agent,0)+sum(int(ir["sure"]) for ir in sistem_invis)

        gecis_f=filter_call(gecis_rows,agent)
        invis_f=invis_clean[~invis_clean["start_ts"].apply(lambda t: in_cw(t,agent) or bfc(t,agent))].copy() if not invis_clean.empty else invis_clean

        tekil_gecis,real_inv=merge_gecis_invisible(gecis_f,invis_f)
        has_t="Toplantı" in ss_df["statu"].values
        has_call_t=has_call_transition(agent,shift_start,shift_end,cw)
        max_g=4+(1 if has_t else 0)+(1 if has_call_t else 0)

        tg=len(tekil_gecis)+len(real_inv)
        if tg>max_g:
            bonuslar=[]
            if has_t: bonuslar.append("Toplantı +1")
            if has_call_t: bonuslar.append("Call +1")
            bonus_str=f" ({', '.join(bonuslar)})" if bonuslar else ""
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Fazla Geçiş Kullanımı",
                "Detay":f"Toplam Geçiş: {tg} kez / Limit: {max_g}{bonus_str}",
                "İhlal Süresi (sn)":0})

        for gr in tekil_gecis:
            sure=gr["sure"]
            if sure>GECIS_LIMIT:
                ex=sure-GECIS_LIMIT
                violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                    "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                    "İhlal Türü":"Uzun Geçiş Süresi",
                    "Detay":f"Mola/Yemek/Çıkış Geçiş: {gr['start_ts'].strftime('%H:%M:%S')}, Süre: {int(sure//60)}dk {int(sure%60)}sn / Limit: 10dk / Fazla: {int(ex//60)}dk {int(ex%60)}sn",
                    "İhlal Süresi (sn)":int(ex)})

        for ir in real_inv:
            sure=ir["sure"]; ex=max(0,int(sure)-GECIS_LIMIT)
            detay=f"Invisible: {ir['start_ts'].strftime('%H:%M:%S')}, Süre: {int(sure//60)}dk {int(sure%60)}sn"
            if sure>GECIS_LIMIT: detay+=f" / 10dk Limiti Aşıldı, Fazla: {int(ex//60)}dk {int(ex%60)}sn"
            violations.append({"Temsilci":agent,"Tarih":shift_date.strftime("%d.%m.%Y"),
                "Vardiya":f"{shift_start.strftime('%H:%M')}-{shift_end.strftime('%H:%M')}",
                "İhlal Türü":"Hatalı Statü (Invisible)","Detay":detay,"İhlal Süresi (sn)":int(sure)})

    df_v=pd.DataFrame(violations).sort_values(["Temsilci","Tarih"]) if violations else \
         pd.DataFrame(columns=["Temsilci","Tarih","Vardiya","İhlal Türü","Detay","İhlal Süresi (sn)"])
    ew={}
    for agent in df_v["Temsilci"].unique():
        ag_df=df_v[df_v["Temsilci"]==agent]
        msn=ag_df[ag_df["İhlal Türü"].isin(["Geç Giriş","Erken Çıkış"])]["İhlal Süresi (sn)"].sum()
        ew[agent]=calc_tamamlatilan_chat(agent,ss,bd,df_cs,msn)
    login_data={}
    for ag in sorted(set(k[0] for k in ss.keys())):
        chat_t=calc_login_chat(ag,df_cs,ss)
        call_t=0
        for (a,sd),(sh_s,sh_e) in ss.items():
            if a!=ag: continue
            for (cs,ce) in cw.get(ag,[]):
                call_t+=max(0,(min(ce,sh_e)-max(cs,sh_s)).total_seconds())
        login_data[ag]=int(chat_t+call_t)
    return df_v,ew,login_data,sistem_kes,ss

# ── EXCEL ÇIKTI ───────────────────────────────────────────────────────────────
def build_excel_bytes(df_v, ew, login_data, ihlal_types, sistem_kes=None, all_agents=None):
    wb=Workbook(); ws=wb.active; ws.title="Vardiya İhlalleri"
    hf=PatternFill("solid",fgColor="2F4F8F"); hfont=Font(bold=True,color="FFFFFF",size=11)
    thin=Side(style="thin",color="CCCCCC"); brd=Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci,h in enumerate(["Temsilci","Tarih","Vardiya","İhlal Türü","Detay","İhlal Süresi"],1):
        c=ws.cell(row=1,column=ci,value=h); c.fill=hf; c.font=hfont
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
    ws.row_dimensions[1].height=25
    for ri,(_,row) in enumerate(df_v.iterrows(),2):
        fill=PatternFill("solid",fgColor=COLORS.get(row["İhlal Türü"],"FFFFFF"))
        for ci,val in enumerate([row["Temsilci"],row["Tarih"],row["Vardiya"],row["İhlal Türü"],row["Detay"],fmt(row["İhlal Süresi (sn)"])],1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=fill
            c.alignment=Alignment(horizontal="left" if ci==5 else "center",vertical="center",wrap_text=True); c.border=brd
        ws.row_dimensions[ri].height=30
    for col,w in zip(["A","B","C","D","E","F"],[25,13,15,25,70,18]): ws.column_dimensions[col].width=w

    ws2=wb.create_sheet("Özet")
    has_kes=sistem_kes is not None
    sum_headers=(["Temsilci"]+ihlal_types+
                 ["Toplam İhlal","Toplam İhlal Süresi","Tamamlatılan Süre","Login Süresi"]+
                 (["Sistem Kesintisi"] if has_kes else []))
    for ci,h in enumerate(sum_headers,1):
        c=ws2.cell(row=1,column=ci,value=h); c.fill=hf; c.font=hfont
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=brd
    ws2.row_dimensions[1].height=30
    kes_fill=PatternFill("solid",fgColor="FFE0CC")
    agents=all_agents if all_agents else sorted(df_v["Temsilci"].unique())
    for ri,agent in enumerate(agents,2):
        ag_df=df_v[df_v["Temsilci"]==agent] if len(df_v)>0 else pd.DataFrame()
        ag=ag_df["İhlal Türü"].value_counts() if len(ag_df)>0 else pd.Series()
        tsn=ag_df["İhlal Süresi (sn)"].sum() if len(ag_df)>0 else 0
        esn=ew.get(agent,0); ksn=(sistem_kes or {}).get(agent,0)
        counts=[ag.get(t,0) for t in ihlal_types]
        row_data=([agent]+counts+[sum(counts),fmt(tsn),fmt(esn),fmt(login_data.get(agent,0))]+
                  ([fmt(ksn) if ksn>0 else "-"] if has_kes else []))
        for ci,val in enumerate(row_data,1):
            c=ws2.cell(row=ri,column=ci,value=val)
            c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
            if ci>=len(ihlal_types)+2: c.font=Font(bold=True)
            if has_kes and ci==len(sum_headers) and ksn>0: c.fill=kes_fill
        ws2.row_dimensions[ri].height=18
    total_row=len(agents)+2
    tsa=df_v["İhlal Süresi (sn)"].sum() if len(df_v)>0 else 0
    tkes=sum((sistem_kes or {}).values())
    td=(["GENEL TOPLAM"]+
        [df_v[df_v["İhlal Türü"]==t].shape[0] if len(df_v)>0 else 0 for t in ihlal_types]+
        [len(df_v),fmt(tsa),fmt(sum(ew.values())),fmt(sum(login_data.values()))]+
        ([fmt(tkes) if tkes>0 else "-"] if has_kes else []))
    tf=PatternFill("solid",fgColor="1F3864"); tft=Font(bold=True,color="FFFFFF")
    for ci,val in enumerate(td,1):
        c=ws2.cell(row=total_row,column=ci,value=val); c.fill=tf; c.font=tft
        c.alignment=Alignment(horizontal="center",vertical="center"); c.border=brd
    ws2.row_dimensions[total_row].height=22
    ws2.column_dimensions["A"].width=25
    for col in "BCDEFGHIJK": ws2.column_dimensions[col].width=19
    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
